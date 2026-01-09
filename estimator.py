# app.py
# Streamlit app that uses ONLY the "Maintenance Estimate" sheet in your Excel model
# as the calculation engine (writes inputs -> recalculates formulas -> reads totals).

import streamlit as st
import openpyxl
import calendar
import tempfile
import shutil
from xlcalculator import ModelCompiler, Evaluator

# ------------------ CONFIG ------------------
MODEL_PATH = "estimate_model.xlsx"  # rename your uploaded file to this and place it next to app.py
SHEET_NAME = "Maintenance Estimate"

# 🔧 Update these cell references to match your sheet:
INPUT_CELLS = {
    "project": "C2",
    "address": "C3",
    "sqft": "C6",
}

# 🔧 Update these output cells to match where totals live ON THE SAME SHEET:
OUTPUT_CELLS = {
    "basic_total": "J22",
    "gold_total": "J23",
    "platinum_total": "J24",
}


# ------------------ ENGINE ------------------
def calculate_maintenance_estimate(inputs: dict) -> dict:
    """
    Writes inputs into Maintenance Estimate sheet, recalculates formulas,
    then returns totals read from Maintenance Estimate sheet only.
    """

    # Work on a temp copy so your source file never gets altered/corrupted
    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    shutil.copy(MODEL_PATH, tmp.name)

    # 1) Write inputs into the Excel sheet
    wb = openpyxl.load_workbook(tmp.name)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f'Sheet "{SHEET_NAME}" not found. Found: {wb.sheetnames}')

    ws = wb[SHEET_NAME]

    ws[INPUT_CELLS["project"]] = inputs.get("project", "")
    ws[INPUT_CELLS["address"]] = inputs.get("address", "")
    ws[INPUT_CELLS["sqft"]] = float(inputs.get("sqft", 0))

    wb.save(tmp.name)

    # 2) Recalculate formulas (Excel calculation engine via xlcalculator)
    compiler = ModelCompiler()
    model = compiler.read_and_parse_archive(tmp.name)
    evaluator = Evaluator(model)

    # 3) Read outputs from the SAME sheet only
    results = {}
    for key, cell in OUTPUT_CELLS.items():
        # xlcalculator expects "SheetName!A1"
        results[key] = evaluator.evaluate(f"{SHEET_NAME}!{cell}")

    return results


# ------------------ STREAMLIT UI ------------------
st.set_page_config(page_title="Maintenance Estimate", layout="centered")
st.title("Landscape Maintenance Estimate")

with st.form("estimate_form"):
    project = st.text_input("Client / Project Name", value="")
    address = st.text_input("Property Address", value="")
    sqft = st.number_input("Property Square Footage", min_value=1000, step=500, value=19000)

    submitted = st.form_submit_button("Generate Estimate")

if submitted:
    try:
        with st.spinner("Calculating..."):
            results = calculate_maintenance_estimate(
                {"project": project, "address": address, "sqft": sqft}
            )

        st.subheader("Estimated Annual Pricing")
        st.metric("Basic Package", f"${float(results['basic_total']):,.2f}")
        st.metric("Gold Package", f"${float(results['gold_total']):,.2f}")
        st.metric("Platinum Package", f"${float(results['platinum_total']):,.2f}")

        st.caption(
            "Note: This app uses your Excel sheet as the calculation engine. "
            "If you change formulas/pricing in the spreadsheet, the app updates automatically."
        )

    except Exception as e:
        st.error(f"Error calculating estimate: {e}")

st.divider()
st.caption(
    "Setup: Place 'estimate_model.xlsx' in the same folder as this file and run:\n"
    "pip install streamlit openpyxl xlcalculator\n"
    "streamlit run app.py\n"
)
